import os
import re
import json
import yaml
import logging
from datetime import datetime

from pipeline_utils import (
    bootstrap_system_paths,
    PipelineIndentedFormatter,
    save_json_file,
    setup_root_console_logging,
    CONFIG_FILE,
)

# =============================================================================
# SECTION 1: BASE SCRAPER CLASS
# =============================================================================
# This module defines the abstract foundation that every spreadsheet scraper
# inherits from. It handles the cross-cutting concerns that all scrapers share:
#   - YAML configuration loading and path resolution
#   - Per-spreadsheet logger creation with isolated log files
#   - State-file integration (detecting whether a re-scrape is required)
#   - Data sanitization (cleaning perk cells, weapon names, info fields)
#   - File I/O abstractions (CSV vs XLSX, row-as-dicts vs raw rows)
#
# Concrete subclasses must implement _execute_processing(), which is the
# entry point invoked by run().
class BaseSpreadsheetScraper:
    def __init__(self, spreadsheet_key):
        """
        Initialize the scraper for a specific spreadsheet key.

        spreadsheet_key is the logical identifier used in config.yaml and in
        the state file (e.g., "pve_rolls", "pvp_rolls"). It also becomes the
        logger name and the prefix for output filenames.
        """
        # Ensure root console logging is alive so manual runs emit to stdout.
        setup_root_console_logging()

        self.spreadsheet_key = spreadsheet_key

        # Load pipeline-wide configuration (paths, API keys, filtering rules).
        self.config = self._load_pipeline_configuration()

        # Resolve directories. Defaults are safety nets in case config.yaml is
        # missing or does not contain the pipeline_paths stanza.
        paths_config = self.config.get("pipeline_paths", {})

        self.source_dir = paths_config.get("download_dir", "workbooks_downloaded")
        self.logs_dir = paths_config.get("log_dir", "logs")
        self.state_file = paths_config.get("state_file", "local_version_state.json")
        self.output_dir = paths_config.get("scraped_dir", "workbooks_scraped_data")

        # Create directories eagerly so loggers and downstream writers never
        # hit FileNotFoundError on first run.
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.source_dir, exist_ok=True)
        os.makedirs(self.logs_dir, exist_ok=True)

        # Spin up the isolated logger before any other work so every step below
        # is captured to the scraper's own log file.
        self._initialize_logging_system()

        # Emit a warning if config.yaml could not be loaded; this helps debug
        # path issues without crashing the pipeline.
        self._report_configuration_status()

    def _initialize_logging_system(self):
        """
        Create a dedicated logger named after the spreadsheet key.

        Each scraper gets its own .log file so that when the scheduler runs
        multiple scrapers back-to-back, their output does not interleave in a
        single shared file.
        """
        log_file_path = os.path.join(self.logs_dir, f"{self.spreadsheet_key}_scraper.log")

        # Named loggers allow fine-grained control per scraper if the pipeline
        # ever needs to suppress or elevate one sheet independently.
        self.logger = logging.getLogger(self.spreadsheet_key)
        self.logger.setLevel(logging.INFO)

        # Defensive reset: in long-running processes or notebook environments,
        # re-instantiating the same scraper would duplicate handlers.
        if self.logger.hasHandlers():
            self.logger.handlers.clear()

        # PipelineIndentedFormatter preserves multi-line messages with visual
        # indentation so stack traces and perk lists remain readable.
        LOG_LAYOUT = "%(asctime)s [%(levelname)s] [%(name)s] -> %(message)s"
        custom_formatter = PipelineIndentedFormatter(fmt=LOG_LAYOUT)

        file_handler = logging.FileHandler(log_file_path, encoding='utf-8')
        file_handler.setFormatter(custom_formatter)
        self.logger.addHandler(file_handler)

        # Visual banner so a human reading the log can instantly see where a
        # new scraper run begins.
        self.logger.info(f"================================================================================")
        self.logger.info(f"🚀 Initializing Data Extraction Pipeline Engine: [{self.spreadsheet_key}]")
        self.logger.info(f"================================================================================")

    def _load_pipeline_configuration(self):
        """
        Load config.yaml from the canonical CONFIG_FILE path.

        Returns an empty dict on any failure so the scraper can degrade
        gracefully with hardcoded defaults rather than crashing.
        """
        config_path = CONFIG_FILE
        self._config_error = None
        self._config_missing = False

        if os.path.exists(config_path):
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    return yaml.safe_load(f) or {}
            except Exception as e:
                # YAML parse errors or permission issues are stored for later
                # reporting rather than raising, because some scrapers may not
                # need config at all (e.g., they only use hardcoded paths).
                self._config_error = e
                return {}

        # If the file is absent entirely, flag it so _report_configuration_status
        # can warn the operator.
        self._config_missing = True
        return {}

    def _report_configuration_status(self):
        """
        Emit a single log line about config health.

        Called once during __init__ so the warning appears at the top of the
        log file and is not buried under hundreds of row-processing lines.
        """
        if self._config_error:
            self.logger.error(f"❌ Failed to parse config.yaml: {self._config_error}. Falling back to default paths.")
        elif self._config_missing:
            self.logger.warning(f"⚠️ Master configuration data asset missing at config.yaml. Falling back to default paths.")

    # =========================================================================
    # DATA SANITIZATION
    # =========================================================================
    # The spreadsheets are community-curated Google Sheets. Cells contain
    # human inconsistencies: extra newlines, "Version 1.2" annotations mixed
    # into weapon names, "N/A", "FIXED", "/", "-" placeholder values, and
    # BOM characters exported by Excel. The methods below normalize these
    # into a strict internal schema so downstream converters never have to
    # deal with raw-sheet noise.

    def parse_name_and_version(self, raw_name):
        """
        Split a raw weapon-name cell into (display_name, version_string).

        Community sheets often store the version on the second line of the
        cell (e.g., "Palindrome\nVersion 3.0.1"). We strip the BOM, drop
        empty lines, and remove the literal word "version" from the second
        line so the converter gets a clean semver-ish string.
        """
        if not raw_name:
            return "", ""

        # Split on newlines, strip whitespace, discard blank lines.
        name_lines = [line.strip() for line in str(raw_name).split('\n') if line.strip()]
        if not name_lines:
            return "", ""

        # Remove zero-width no-break space (BOM / \ufeff) that Excel and
        # Google Sheets sometimes prepend when exporting to CSV.
        clean_name = name_lines[0].replace("\ufeff", "").strip()
        version_string = ""

        # If there is a second line, treat it as the version annotation.
        if len(name_lines) > 1:
            # Case-insensitive removal of the word "version" plus surrounding
            # whitespace so "Version 1.2" becomes "1.2".
            version_string = re.sub(r'(?i)\s*version\s*', '', name_lines[1]).strip()

        return clean_name, version_string

    def sanitize_perk_cell(self, cell_value):
        """
        Convert a raw perk cell into a list of clean perk names.

        Perk cells are newline-delimited. We drop lines that are purely
        structural placeholders ("/", "N/A", "FIXED", "NONE", "-") because
        they carry no semantic value for the wishlist generator.
        """
        if cell_value is None:
            return []

        lines = [line.strip() for line in str(cell_value).split('\n') if line.strip()]
        # Filter out known noise tokens. We uppercase the comparison so
        # "n/a", "N/a", and "fixed" are all caught.
        return [l for l in lines if l.upper() not in ["/", "N/A", "FIXED", "NONE", "-"]]

    def sanitize_info_cell(self, key, cell_value):
        """
        Normalize an info/metadata cell (rank, tier, role, notes, etc.).

        Rules:
          - None -> empty string.
          - "NONE", "N/A", "-" are collapsed to empty string EXCEPT for the
            "rank" column, where those values may be meaningful (e.g., a weapon
            that genuinely has no PvE rank yet).
          - Multi-line cells are joined with ", " so downstream JSON stays
            single-line and diff-friendly.
        """
        if cell_value is None:
            return ""

        val = str(cell_value).strip()
        # Only collapse placeholder tokens if the column is NOT "rank".
        if str(key).lower() != "rank" and val.upper() in ["NONE", "N/A", "-"]:
            return ""

        # Flatten newlines into comma-separated prose. This prevents the
        # processed JSON from containing embedded \n that break DIM's parser.
        lines = [line.strip() for line in val.split('\n') if line.strip()]
        return ", ".join(lines)

    # =========================================================================
    # STATE REGISTRY
    # =========================================================================
    # The pipeline uses a JSON state file to avoid redundant work. After a
    # successful scrape, the scraper resets the workbook's update flag. The
    # downloader (or a manual trigger) sets the flag when it detects that the
    # Google Sheet has changed. is_update_required() is the gatekeeper.

    def is_update_required(self, workbook_name):
        """
        Determine whether this workbook needs to be re-scraped.

        Three conditions force a True (scrape required):
          1. The source file (CSV/XLSX) is missing -> False, because there is
             nothing to scrape; the downloader should have produced it.
          2. The expected JSON output is missing -> True, because even if the
             state flag is clear, the artifact is absent (deleted, first run).
          3. The state file explicitly flags workbook_scrape_update_required.
        """
        # Guard: if the downloader has not produced the workbook yet, do not
        # attempt to scrape. The scheduler will retry on the next cycle.
        if not os.path.exists(self.get_workbook_file_path(workbook_name)):
            self.logger.warning(f"⚠️ Expected workbook storage asset not found: {workbook_name}")
            return False

        # Derive the expected output path. Subclasses may override
        # output_filename; if not, we synthesize one from the spreadsheet key
        # and a sanitized workbook name.
        expected_output_path = getattr(self, 'output_filename', None)
        if not expected_output_path:
            clean_wb_name = str(workbook_name).lower().replace(" ", "_")
            expected_output_path = os.path.join(self.output_dir, f"{self.spreadsheet_key}_{clean_wb_name}.json")

        # If the JSON artifact is gone, force a scrape regardless of flags.
        # This handles manual deletion, git clean, or disk corruption.
        if expected_output_path and not os.path.exists(expected_output_path):
            self.logger.info(f"🔄 Output data asset missing at '{expected_output_path}'. Forcing scrape routine override.")
            return True

        # Consult the state file. If it is unreadable, default to True as a
        # safe fallback; a bad state file should never cause data staleness.
        if os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as f:
                    state = json.load(f)

                sheet_entry = state.get("spreadsheets", {}).get(self.spreadsheet_key, {})
                workbook_entry = sheet_entry.get("workbooks", {}).get(workbook_name, {})

                return workbook_entry.get("workbook_scrape_update_required", False)
            except Exception as e:
                self.logger.error(f"❌ State file corrupt or unreadable: {e}. Defaulting to full compile safety block.")
                return True

        # No state file at all implies a first run; scrape everything.
        return True

    def reset_scraper_flag(self, workbooks_list):
        """
        Clear the workbook_scrape_update_required flag for every workbook in
        workbooks_list inside the shared state file.

        Called by subclasses after a successful scrape so the next pipeline
        cycle skips these workbooks unless the downloader detects a new version.
        """
        if os.path.exists(self.state_file):
            try:
                with open(self.state_file, "r", encoding="utf-8") as f:
                    state = json.load(f)

                # Navigate to the nested dict: spreadsheets -> <key> -> workbooks
                sheet_entry = state.get("spreadsheets", {}).get(self.spreadsheet_key, {})
                workbooks_entry = sheet_entry.get("workbooks", {})

                for wb in workbooks_list:
                    if wb in workbooks_entry:
                        workbooks_entry[wb]["workbook_scrape_update_required"] = False

                # Atomic(ish) write via pipeline_utils helper.
                save_json_file(self.state_file, state)

                self.logger.info(f"✅ Registry synchronization update flags reset to False for elements: {workbooks_list}")
            except Exception as e:
                # If we cannot write the state file, the next run will re-scrape
                # redundantly. That is harmless but wastes time, so we log it
                # at CRITICAL to alert the operator.
                self.logger.critical(f"❌ Failed to reset local state update flags: {e}")

    # =========================================================================
    # FILE IO UTILITIES
    # =========================================================================
    # These helpers abstract away the difference between CSV and XLSX sources
    # so that subclasses can focus on schema logic rather than parsing.

    def initialize_unified_record(self):
        """
        Return a fresh dict matching the canonical scraped-record schema.

        Every weapon roll record produced by a subclass should conform to this
        shape so the wishlist converter can consume it blindly. The schema
        separates:
          - version       -> semver string from parse_name_and_version.
          - perks         -> 5 columns: column1, column2, perk1, perk2, origin_trait.
          - info          -> metadata (rank, tier, tags, source, notes, etc.).
          - info.analysis -> 8 scoring dimensions used by DIM's roll grading.
        """
        return {
            "version": "",
            "perks": {
                "column1": [], "column2": [], "perk1": [], "perk2": [], "origin_trait": []
            },
            "info": {
                "rank": "", "tier": "", "priority": "", "role": "", "purpose": "", "tags": "", "type": "",
                "usage": "", "source": "", "notes": "", "description": "", "alternatives": "",
                "analysis": {
                    "roam": "", "dps": "", "day1": "", "chall": "", "speed": "", "effect": "", "flex": "", "power": ""
                }
            }
        }

    def _write_output_payload(self, payload, workbooks_dict, label="payload"):
        """
        Write the fully scraped payload to self.output_filename.

        Safety check: if the payload contains zero items across all workbooks,
        we abort the write. Writing an empty file would overwrite a previous
        good run with nothing, which is almost always a bug (missing openpyxl,
        locked XLSX, empty source sheet, etc.).
        """
        total_items = sum(len(v) for v in workbooks_dict.values() if isinstance(v, dict))
        if total_items == 0:
            self.logger.critical("❌ All workbooks produced zero items. Aborting write to prevent data loss.")
            self.logger.critical("   Possible causes: missing 'openpyxl', locked file, corrupted download, or empty source sheet.")
            return False

        with open(self.output_filename, "w", encoding="utf-8") as f_out:
            json.dump(payload, f_out, indent=2, ensure_ascii=False)
        self.logger.info(f"🎉 Successfully compiled {label} down to cache: {self.output_filename}")
        return True

    def _get_workbook_config(self, workbook_name):
        """
        Look up the workbook-specific stanza inside config.yaml.

        Returns an empty dict if the workbook is not configured, which lets
        the caller fall back to default behavior (e.g., CSV instead of XLSX).
        """
        ss_config = self.config.get("spreadsheets", {}).get(self.spreadsheet_key, {})
        for wb in ss_config.get("workbooks", []):
            if wb.get("name") == workbook_name:
                return wb
        return {}

    def get_workbook_file_path(self, workbook_name):
        """
        Resolve the on-disk path for a given workbook.

        The filename pattern is: <spreadsheet_key>_<sanitized_workbook_name>.<<ext>
        where ext is "xlsx" if the config explicitly sets use_xlsx, otherwise
        "csv". This convention lets multiple scrapers share the same source_dir
        without collisions.
        """
        clean_wb_name = str(workbook_name).lower().replace(" ", "_")
        wb_config = self._get_workbook_config(workbook_name)

        ext = "xlsx" if wb_config.get("use_xlsx") else "csv"
        return os.path.join(self.source_dir, f"{self.spreadsheet_key}_{clean_wb_name}.{ext}")

    def _read_file_as_dicts(self, file_path, skip_rows=0):
        """
        Parse a CSV or XLSX file into a list of dicts, one per data row.

        skip_rows allows skipping header rows or banner rows that community
        sheets sometimes insert above the real column headers.

        For XLSX:
          - We use openpyxl with data_only=True so formulas are evaluated to
            their cached values rather than returning formula strings.
          - The first non-skipped row becomes the header keys.
          - Each subsequent row becomes a dict keyed by those headers.

        For CSV:
          - We skip N rows manually, then hand the stream to csv.DictReader.
          - Keys and values are stripped of surrounding whitespace.
        """
        if not os.path.exists(file_path):
            self.logger.error(f"❌ Cannot ingest dictionary records. Path missing: {file_path}")
            return []

        if file_path.endswith(".xlsx"):
            try:
                import openpyxl
            except ImportError:
                self.logger.critical("❌ Required execution library module dependency 'openpyxl' is missing! Run pip install openpyxl.")
                return []

            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            raw_rows = list(sheet.iter_rows(values_only=True))
            if len(raw_rows) <= skip_rows:
                return []

            # Discard the skip_rows banner/header padding.
            raw_rows = raw_rows[skip_rows:]
            # First surviving row is the true header.
            header_row = [str(cell).strip() if cell is not None else None for cell in raw_rows[0]]

            dict_rows = []
            for row in raw_rows[1:]:
                row_dict = {}
                for idx, col_name in enumerate(header_row):
                    if col_name is not None:
                        val = row[idx] if idx < len(row) else None
                        row_dict[col_name] = val
                dict_rows.append(row_dict)
            return dict_rows
        else:
            import csv
            with open(file_path, mode="r", encoding="utf-8-sig") as f:
                # utf-8-sig strips the BOM that Excel-injected CSVs carry.
                for _ in range(skip_rows):
                    next(f, None)
                reader = csv.DictReader(f)
                # Strip whitespace from keys so "perk 1 " and "perk 1" do not
                # create divergent dict keys across different sheets.
                return [{k.strip() if k else "": v for k, v in row.items()} for row in reader]

    def _read_file_raw_rows(self, file_path):
        """
        Parse a CSV or XLSX file into a list of raw string rows.

        Unlike _read_file_as_dicts, this does not treat any row as a header.
        It returns a 2D string matrix. Used by subclasses that need to parse
        non-standard layouts (e.g., perk columns that are not labeled).
        """
        if not os.path.exists(file_path):
            return []

        if file_path.endswith(".xlsx"):
            try:
                import openpyxl
            except ImportError:
                self.logger.critical("❌ Required execution library module dependency 'openpyxl' is missing!")
                return []
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            # Coerce every cell to a stripped string; None becomes "".
            return [[str(cell).strip() if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        else:
            import csv
            with open(file_path, mode="r", encoding="utf-8-sig") as f:
                return list(csv.reader(f))

    def run(self):
        """
        Public entry point. Wraps _execute_processing() in a try/except that
        logs the full traceback and then re-raises so the scheduler can decide
        whether to retry or abort the entire pipeline.

        Subclasses must implement _execute_processing().
        """
        try:
            self._execute_processing()
        except Exception as err:
            # exc_info=True injects the full traceback into the log file.
            # We re-raise so the caller (scheduler / __main__ block) knows
            # the scraper failed and can set appropriate retry flags.
            self.logger.error(f"❌ Pipeline processing failed during runtime: {str(err)}", exc_info=True)
            raise err